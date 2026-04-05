"""
전기기사 문제 검토 오류 목록을 Excel(.xlsx)로 내보내기
- 년도/회차/과목/원본문제번호 포함
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# ── 스타일 정의 ──
header_font = Font(name='맑은 고딕', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

cell_font = Font(name='맑은 고딕', size=10)
cell_align = Alignment(vertical='top', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

wrong_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')    # 노란색 - 오답
outdated_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')  # 주황색 - 규정변경
ocr_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')       # 파란색 - OCR오류


def style_header(ws, cols):
    for col_idx, title in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def style_row(ws, row_idx, num_cols, fill=None):
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = cell_font
        cell.alignment = center_align if col_idx <= 5 else cell_align
        cell.border = thin_border
        if fill:
            cell.fill = fill


# ── 원본 시험 문제번호 계산 ──
# 전기기사 과목 순서: 전기자기학(1~20), 전력공학(21~40), 전기기기(41~60), 회로이론및제어공학(61~80), 전기설비기술기준(81~100)
SUBJECT_OFFSET = {
    '전기자기학': 0,
    '전력공학': 20,
    '전기기기': 40,
    '회로이론 및 제어공학': 60,
    '전기설비기술기준': 80,
}

def calc_original_num(subject, num_in_subject):
    """과목명 + 과목 내 순번 → 원본 시험지 문제 번호"""
    offset = SUBJECT_OFFSET.get(subject, 0)
    return offset + num_in_subject


# ── DB에서 조회한 원본 정보 ──
# (문제코드, 년도, 회차, 과목, 과목내순번) → 원본번호 계산
QUESTION_INFO = {
    'ELEC-E-CC-016':  ('2018', '1회', '회로이론 및 제어공학', 16),
    'ELEC-E-CC-109':  ('2019', '3회', '회로이론 및 제어공학', 9),
    'ELEC-E-CC-158':  ('2020', '3회', '회로이론 및 제어공학', 18),
    'ELEC-E-CC-251':  ('2022', '1회', '회로이론 및 제어공학', 11),
    'ELEC-E-CC-297':  ('2025', '3회', '회로이론 및 제어공학', 17),
    'ELEC-E-EL-016':  ('2018', '1회', '전기설비기술기준', 16),
    'ELEC-E-EL-036':  ('2018', '2회', '전기설비기술기준', 16),
    'ELEC-E-EL-082':  ('2019', '2회', '전기설비기술기준', 2),
    'ELEC-E-EL-106':  ('2019', '3회', '전기설비기술기준', 6),
    'ELEC-E-EM-007':  ('2018', '1회', '전기자기학', 7),
    'ELEC-E-EM-037':  ('2018', '2회', '전기자기학', 17),
    'ELEC-E-EM-080':  ('2019', '1회', '전기자기학', 20),
    'ELEC-E-EM-170':  ('2020', '4회', '전기자기학', 10),
    'ELEC-E-EM-211':  ('2021', '2회', '전기자기학', 11),
    'ELEC-E-MA-351':  ('2025', '2회', '전기기기', 11),
    'ELEC-E-MA-359':  ('2025', '2회', '전기기기', 19),
    'ELEC-E-XX-032':  ('2018', '1회', '전기기기', 12),
    'ELEC-E-XX-057':  ('2018', '2회', '전력공학', 17),
    'ELEC-E-XX-143':  ('2019', '1회', '전기기기', 3),
    'ELEC-E-XX-196':  ('2019', '2회', '전기기기', 16),
    'ELEC-E-XX-244':  ('2020', '1회', '전력공학', 4),
    'ELEC-E-XX-255':  ('2020', '1회', '전력공학', 15),
    'ELEC-E-XX-350':  ('2020', '4회', '전기기기', 10),
    'ELEC-E-XX-418':  ('2021', '2회', '전력공학', 18),
    'ELEC-E-XX-450':  ('2021', '3회', '전력공학', 10),
}


def get_info(code):
    """문제코드 → (년도, 회차, 과목, 원본번호) 반환"""
    info = QUESTION_INFO.get(code)
    if not info:
        return ('?', '?', '?', '?')
    year, rnd, subj, num_in_subj = info
    orig_num = calc_original_num(subj, num_in_subj)
    return (f'{year}년', rnd, subj, orig_num)


# ═══════════════════════════════════════════
# Sheet 1: 오답 (Wrong Answers)
# ═══════════════════════════════════════════
ws1 = wb.active
ws1.title = '오답 (14건)'

cols1 = ['No', '문제코드', '년도', '회차', '원본 문제번호', '과목', 'DB정답', '올바른 정답', '오류 설명']
style_header(ws1, cols1)

wrong_answers = [
    ('ELEC-E-CC-016', 4, 1, '직렬 RLC 공진 시 임피던스 최소 → 전류 최대가 정답(①). DB의 ④번은 오답.'),
    ('ELEC-E-CC-109', 3, 2, '특성방정식 s²+4s+3=0의 근이 s=-1,-3 → 과감쇠(overdamped). ②번이 정답.'),
    ('ELEC-E-CC-158', 1, 3, 'Y-Δ 변환 공식 적용 시 Zab = 10Ω이 아닌 다른 값. ③번이 정답.'),
    ('ELEC-E-CC-251', 2, 4, '전달함수 계산 결과 ④번이 맞음. DB의 ②번은 계산 오류.'),
    ('ELEC-E-CC-297', 3, 1, '라플라스 역변환 결과 ①번 e^(-t)·sin(2t)가 정답. DB의 ③번은 오답.'),
    ('ELEC-E-EL-106', 4, 2, 'B종 접지저항 = 150/Ig. 1선 지락전류 5A 기준 → 30Ω. ②번이 정답.'),
    ('ELEC-E-EM-007', 2, 4, '평행판 콘덴서 에너지 = εE²V/2 공식에서 ④번이 정답.'),
    ('ELEC-E-EM-037', 1, 3, '환상 솔레노이드 자계 H = NI/(2πr). 계산 결과 ③번이 정답.'),
    ('ELEC-E-EM-080', 3, 1, '정전용량 합성 계산 결과 ①번이 정답. DB의 ③번은 오답.'),
    ('ELEC-E-EM-170', 4, 2, '전자유도 법칙 적용 시 ②번이 정답.'),
    ('ELEC-E-EM-211', 2, 3, '자기회로 자속 계산 결과 ③번이 정답.'),
    ('ELEC-E-XX-255', 3, 4, '변압기 효율 최대 조건: 철손=동손. 계산 결과 ④번이 정답.'),
    ('ELEC-E-XX-350', 1, 2, '송전선 충전전류 계산 결과 ②번이 정답.'),
    ('ELEC-E-XX-418', 4, 1, '차단기 차단용량 계산 시 ①번이 정답. DB의 ④번은 계산 오류.'),
]

for i, (code, db_ans, correct_ans, desc) in enumerate(wrong_answers, 1):
    row = i + 1
    year, rnd, subj, orig_num = get_info(code)
    ws1.cell(row=row, column=1, value=i)
    ws1.cell(row=row, column=2, value=code)
    ws1.cell(row=row, column=3, value=year)
    ws1.cell(row=row, column=4, value=rnd)
    ws1.cell(row=row, column=5, value=orig_num)
    ws1.cell(row=row, column=6, value=subj)
    ws1.cell(row=row, column=7, value=db_ans)
    ws1.cell(row=row, column=8, value=correct_ans)
    ws1.cell(row=row, column=9, value=desc)
    style_row(ws1, row, len(cols1), wrong_fill)

ws1.column_dimensions['A'].width = 5
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 8
ws1.column_dimensions['D'].width = 7
ws1.column_dimensions['E'].width = 13
ws1.column_dimensions['F'].width = 22
ws1.column_dimensions['G'].width = 9
ws1.column_dimensions['H'].width = 11
ws1.column_dimensions['I'].width = 60

# ═══════════════════════════════════════════
# Sheet 2: 규정 변경 (Outdated Regulation)
# ═══════════════════════════════════════════
ws2 = wb.create_sheet('규정변경 (1건)')

cols2 = ['No', '문제코드', '년도', '회차', '원본 문제번호', '과목', 'DB정답', '현행 정답', '변경 내용']
style_header(ws2, cols2)

outdated = [
    ('ELEC-E-EL-036', 1, 2,
     '누전차단기 감도전류 기준이 KEC 2021 시행으로 15mA → 30mA로 변경됨. '
     '기존 정답 ①(15mA)은 구 규정 기준이며, 현행 기준 ②(30mA)이 정답.'),
]

for i, (code, db_ans, correct_ans, desc) in enumerate(outdated, 1):
    row = i + 1
    year, rnd, subj, orig_num = get_info(code)
    ws2.cell(row=row, column=1, value=i)
    ws2.cell(row=row, column=2, value=code)
    ws2.cell(row=row, column=3, value=year)
    ws2.cell(row=row, column=4, value=rnd)
    ws2.cell(row=row, column=5, value=orig_num)
    ws2.cell(row=row, column=6, value=subj)
    ws2.cell(row=row, column=7, value=db_ans)
    ws2.cell(row=row, column=8, value=correct_ans)
    ws2.cell(row=row, column=9, value=desc)
    style_row(ws2, row, len(cols2), outdated_fill)

ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 8
ws2.column_dimensions['D'].width = 7
ws2.column_dimensions['E'].width = 13
ws2.column_dimensions['F'].width = 22
ws2.column_dimensions['G'].width = 9
ws2.column_dimensions['H'].width = 11
ws2.column_dimensions['I'].width = 60

# ═══════════════════════════════════════════
# Sheet 3: OCR/텍스트 오류 (OCR Errors)
# ═══════════════════════════════════════════
ws3 = wb.create_sheet('OCR오류 (10건)')

cols3 = ['No', '문제코드', '년도', '회차', '원본 문제번호', '과목', '오류 위치', '현재 텍스트 (일부)', '수정 필요 내용']
style_header(ws3, cols3)

ocr_errors = [
    ('ELEC-E-EL-016', '보기', '보기 텍스트에 특수문자 깨짐', '특수문자/기호가 올바르게 표시되도록 수정 필요'),
    ('ELEC-E-EL-082', '문제', '수식 일부 누락', '전압 계산 수식에서 첨자/분수 표현 보완 필요'),
    ('ELEC-E-MA-351', '보기', '보기 텍스트 불완전', '보기 3, 4번의 텍스트가 잘려 있어 보완 필요'),
    ('ELEC-E-MA-359', '문제', '수식 렌더링 오류', 'KaTeX 수식 문법 오류로 렌더링 안 됨. 수식 재입력 필요'),
    ('ELEC-E-XX-032', '보기', '보기 번호 매핑 오류', '보기 1~4번 텍스트가 한 칸씩 밀려 있음'),
    ('ELEC-E-XX-057', '문제', '단위 표기 오류', 'kW를 KW로 표기. 올바른 SI 단위(kW)로 수정 필요'),
    ('ELEC-E-XX-143', '보기', '수식 깨짐', '보기의 분수 수식이 텍스트로 깨져 있음. KaTeX로 재입력 필요'),
    ('ELEC-E-XX-196', '문제', '문제 텍스트 중복', '문제 텍스트 일부가 반복 입력되어 있음'),
    ('ELEC-E-XX-244', '보기', '그리스 문자 깨짐', 'ω, θ 등 그리스 문자가 깨져서 표시됨'),
    ('ELEC-E-XX-450', '문제', '줄바꿈 오류', '문제 텍스트에 불필요한 줄바꿈이 삽입되어 가독성 저하'),
]

for i, (code, loc, current, fix) in enumerate(ocr_errors, 1):
    row = i + 1
    year, rnd, subj, orig_num = get_info(code)
    ws3.cell(row=row, column=1, value=i)
    ws3.cell(row=row, column=2, value=code)
    ws3.cell(row=row, column=3, value=year)
    ws3.cell(row=row, column=4, value=rnd)
    ws3.cell(row=row, column=5, value=orig_num)
    ws3.cell(row=row, column=6, value=subj)
    ws3.cell(row=row, column=7, value=loc)
    ws3.cell(row=row, column=8, value=current)
    ws3.cell(row=row, column=9, value=fix)
    style_row(ws3, row, len(cols3), ocr_fill)

ws3.column_dimensions['A'].width = 5
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 8
ws3.column_dimensions['D'].width = 7
ws3.column_dimensions['E'].width = 13
ws3.column_dimensions['F'].width = 22
ws3.column_dimensions['G'].width = 10
ws3.column_dimensions['H'].width = 38
ws3.column_dimensions['I'].width = 48

# ═══════════════════════════════════════════
# Sheet 4: 전체 요약
# ═══════════════════════════════════════════
ws4 = wb.create_sheet('요약')

ws4.cell(row=1, column=1, value='전기기사 1,600문제 검토 결과 요약').font = Font(name='맑은 고딕', bold=True, size=14)
ws4.merge_cells('A1:C1')

summary = [
    ('검토 대상', '전기기사 16개 시험, 1,600문제'),
    ('검토 일자', '2026-04-05'),
    ('', ''),
    ('오류 유형', '건수'),
    ('오답 (DB 정답이 틀림)', '14건'),
    ('규정 변경 (현행법 기준 오답)', '1건'),
    ('OCR/텍스트 오류', '10건'),
    ('', ''),
    ('합계', '25건'),
    ('오류율', '25/1600 = 1.56%'),
]

for i, (label, value) in enumerate(summary, 3):
    ws4.cell(row=i, column=1, value=label).font = Font(name='맑은 고딕', bold=(label in ['오류 유형', '합계', '오류율']), size=11)
    ws4.cell(row=i, column=2, value=value).font = Font(name='맑은 고딕', bold=(label in ['합계', '오류율']), size=11)

ws4.column_dimensions['A'].width = 35
ws4.column_dimensions['B'].width = 35

# ── 저장 ──
output_path = r'D:\Antigravity\electric-jjang\data\전기기사_문제검토_오류목록.xlsx'
wb.save(output_path)
print(f'저장 완료: {output_path}')
