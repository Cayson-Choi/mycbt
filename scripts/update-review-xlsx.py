"""전기기사 필기 문제 검토 오류목록 - 2018~2025 통합"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=11, color='FFFFFF')
header_fill_blue = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_fill_orange = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
header_fill_green = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
done_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # 연한 초록
todo_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')   # 연한 노랑
verify_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')  # 연한 빨강
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def setup_sheet(ws, headers, data, widths, header_fill, action_col=None):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 30
    for i, row_data in enumerate(data, 2):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col >= len(headers) - 1))
        # 조치사항 컬럼 배경색
        if action_col:
            action_val = str(row_data[action_col - 1]) if row_data[action_col - 1] else ''
            if '수정완료' in action_val or '완료' in action_val or '조치불필요' in action_val:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).fill = done_fill
            elif '원본 대조' in action_val or '시험지 확인' in action_val:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).fill = verify_fill
            elif 'DB 정답' in action_val or '수정' in action_val or '확인' in action_val:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).fill = todo_fill
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# 1. 정답 오류 (41건 = 기존 14 + 신규 27)
# ============================================================
# [No, 코드, 년도, 회차, 문제번호, 과목, DB정답, 올바른정답, 오류설명, 조치사항]
answer_errors = [
    # --- 2018년 ---
    [1,  'ELEC-E-CC-016',  '2018년', '1회', 76,     '회로이론 및 제어공학', 4, 1, '직렬 RLC 공진 시 리액턴스 최소 및 전류 최대가 정답(①). DB에 ④로 저장.', 'DB 정답 4→1 수정 필요'],
    [2,  'ELEC-E-EM-007',  '2018년', '1회', 7,      '전기자기학',         2, 4, '평행판 콘덴서 에너지 = εE²V/2 공식에서 정답이 ④. DB에 ②로 저장.', 'DB 정답 2→4 수정 필요'],
    [3,  'ELEC-E-EM-037',  '2018년', '2회', 17,     '전기자기학',         1, 3, '환상 솔레노이드 자계 H = NI/(2πr). 계산 결과 정답 ③. DB에 ①로 저장.', 'DB 정답 1→3 수정 필요'],
    # --- 2019년 ---
    [4,  'ELEC-E-EM-080',  '2019년', '1회', 20,     '전기자기학',         3, 1, '정전용량 합성 계산 결과 정답 ①. DB에 ③으로 저장.', 'DB 정답 3→1 수정 필요'],
    [5,  'ELEC-E-CC-109',  '2019년', '3회', 69,     '회로이론 및 제어공학', 3, 2, '특성방정식 s²+4s+3=0의 근 s=-1,-3 → 과제동(overdamped). 정답 ②. DB에 ③으로 저장.', 'DB 정답 3→2 수정 필요'],
    [6,  'ELEC-E-EL-106',  '2019년', '3회', 86,     '전기설비기술기준',    4, 2, 'B종 접지저항 = 150/Ig. 1선 지락전류 5A 조건 시 30Ω. 정답 ②. DB에 ④로 저장.', 'DB 정답 4→2 수정 필요'],
    # --- 2020년 ---
    [7,  'ELEC-E-CC-158',  '2020년', '3회', 78,     '회로이론 및 제어공학', 1, 3, 'Y-Δ 변환 공식 적용 시 Zab = 10이 아닌 다른 값. 정답 ③. DB에 ①으로 저장.', 'DB 정답 1→3 수정 필요'],
    [8,  'ELEC-E-XX-255',  '2020년', '1회', 35,     '전력공학',           3, 4, '변압기 효율 최대 조건: 철손=동손. 계산 결과 정답 ④. DB에 ③으로 저장.', 'DB 정답 3→4 수정 필요'],
    [9,  'ELEC-E-EM-170',  '2020년', '4회', 10,     '전기자기학',         4, 2, '비오사바르 법칙 적용 시 정답 ②. DB에 ④로 저장.', 'DB 정답 4→2 수정 필요'],
    [10, 'ELEC-E-XX-350',  '2020년', '4회', 50,     '전기기기',           1, 2, '동기기 전기자반작용 계산 결과 정답 ②. DB에 ①으로 저장.', 'DB 정답 1→2 수정 필요'],
    # --- 2021년 ---
    [11, 'ELEC-E-EM-211',  '2021년', '2회', 11,     '전기자기학',         2, 3, '자기회로 자속 계산 결과 정답 ③. DB에 ②로 저장.', 'DB 정답 2→3 수정 필요'],
    [12, 'ELEC-E-XX-418',  '2021년', '2회', 38,     '전력공학',           4, 1, '차단기 차단용량 계산 시 정답 ①. DB에 ④로 저장.', 'DB 정답 4→1 수정 필요'],
    # --- 2022년 ---
    [13, 'ELEC-E-CC-251',  '2022년', '1회', 71,     '회로이론 및 제어공학', 2, 4, '전달함수 최종 값의 정리에서 정답 ④. DB에 ②로 저장.', 'DB 정답 2→4 수정 필요'],
    # --- 2023년 ---
    [14, 'ELEC-E-CC-385',  '2023년', '2회', 'CC-05', '회로이론 및 제어공학', 1, 4, '2전력계법 역률: tanφ=√3(500-300)/800=0.433, cosφ=91.8%. 정답 ④(91.6%)이나 DB에 ①(70.7%)', 'DB 정답 1→4 수정 필요'],
    [15, 'ELEC-E-CC-390',  '2023년', '2회', 'CC-10', '회로이론 및 제어공학', 1, 2, '비정현파 실효값=√(10²+5²)=11.18V. 정답 ②(11.2V)이나 DB에 ①(3.87V)', 'DB 정답 1→2 수정 필요'],
    [16, 'ELEC-E-CC-395',  '2023년', '2회', 'CC-15', '회로이론 및 제어공학', 2, 1, '1/(s²+3s+1): ζ=3/2=1.5>1 → 과제동. 정답 ①이나 DB에 ②(부족제동)', 'DB 정답 2→1 수정 필요'],
    [17, 'ELEC-E-EM-328',  '2023년', '2회', 'EM-08', '전기자기학',         1, 2, '변위전류 i_d=CωVmcosωt (C에 ε 포함). 정답 ②이나 DB에 ①(εCωVmcosωt)', 'DB 정답 1→2 수정 필요'],
    [18, 'ELEC-E-EL-403',  '2023년', '2회', 'EL-03', '전기설비기술기준',    1, 4, '지중전선로 차량 중량물 압력 장소 매설깊이 1.2m. 정답 ④이나 DB에 ①(0.6m)', 'DB 정답 1→4 수정 필요'],
    [19, 'ELEC-E-EL-409',  '2023년', '2회', 'EL-09', '전기설비기술기준',    1, 4, '22.9kV 특고압 가공전선 도로횡단 높이 6m. 정답 ④이나 DB에 ①(4.5m)', 'DB 정답 1→4 수정 필요'],
    [20, 'ELEC-E-CC-415',  '2023년', '3회', 'CC-15', '회로이론 및 제어공학', 3, 1, '감쇠비=제2오버슈트/피크오버슈트(비). 정답 ①(나눈값)이나 DB에 ③(곱한값)', 'DB 정답 3→1 수정 필요'],
    [21, 'ELEC-E-CC-416',  '2023년', '3회', 'CC-16', '회로이론 및 제어공학', 2, 3, 'ζ=1,ωₙ=1 → G(s)=1/(s+1)², 임펄스응답=te^(-t). 정답 ③이나 DB에 ②', 'DB 정답 2→3 수정 필요'],
    # --- 2024년 ---
    [22, 'ELEC-E-CC-432',  '2024년', '1회', 'CC-12', '회로이론 및 제어공학', 2, 1, '논리식 x̄y+ȳx+xy=(0,1,1,1)=x+y. 정답 ①이나 DB에 ②(x̄+y)', 'DB 정답 2→1 수정 필요'],
    [23, 'ELEC-E-CC-434',  '2024년', '1회', 'CC-14', '회로이론 및 제어공학', 4, 2, '대칭좌표법 역상분 I₂=2.51∠96.55°. 정답 ②이나 DB에 ④(정상분 값)', 'DB 정답 4→2 수정 필요'],
    [24, 'ELEC-E-EM-361',  '2024년', '1회', 'EM-01', '전기자기학',         2, 1, '전위차 V=σd/ε₀=1.807×10¹²V. 정답 ①이나 DB에 ②(10¹¹, 10배 오류)', 'DB 정답 2→1 수정 필요'],
    [25, 'ELEC-E-EM-367',  '2024년', '1회', 'EM-07', '전기자기학',         3, 4, '회전력 T=1.44×10⁻³N·m. 정답 ④이나 DB에 ③(10⁻⁵, 100배 오류)', 'DB 정답 3→4 수정 필요'],
    [26, 'ELEC-E-EM-376',  '2024년', '1회', 'EM-16', '전기자기학',         2, 4, 'η=60π → εᵣ=4, v=1.5×10⁸, ω=1.5×10⁸. 정답 ④이나 DB에 ②(3×10⁸)', 'DB 정답 2→4 수정 필요'],
    [27, 'ELEC-E-MA-419',  '2024년', '1회', 'MA-19', '전기기기',           3, 1, '최대전압변동률 εmax=√(p²+q²)=3.07%≈3.1%. 정답 ①이나 DB에 ③(5.1%)', 'DB 정답 3→1 수정 필요'],
    [28, 'ELEC-E-PC-387',  '2024년', '1회', 'PC-07', '전력공학',           1, 2, '단락용량=80MVA/0.3=266.7MVA. 정답 ②이나 DB에 ①(226.7MVA)', 'DB 정답 1→2 수정 필요'],
    [29, 'ELEC-E-CC-441',  '2024년', '2회', 'CC-01', '회로이론 및 제어공학', 2, 3, '비정현파 실효치=√(100²+40²+30²)≈115.7V. 정답 ③이나 DB에 ②(38.6V)', 'DB 정답 2→3 수정 필요'],
    [30, 'ELEC-E-CC-444',  '2024년', '2회', 'CC-04', '회로이론 및 제어공학', 3, 1, 'RLC 소비전력 P=(170×8.5/2)cos60°=361W. 정답 ①이나 DB에 ③(720W)', 'DB 정답 3→1 수정 필요'],
    [31, 'ELEC-E-CC-459',  '2024년', '2회', 'CC-19', '회로이론 및 제어공학', 3, 4, '정상편차 ess=1/Kv=8/5. 정답 ④이나 DB에 ③(5/8, Kv와 혼동)', 'DB 정답 3→4 수정 필요'],
    [32, 'ELEC-E-CC-467',  '2024년', '3회', 'CC-07', '회로이론 및 제어공학', 1, 4, 'AI 검토: 정답 오류 확인', '원본 시험지 확인 후 DB 수정'],
    [33, 'ELEC-E-CC-474',  '2024년', '3회', 'CC-14', '회로이론 및 제어공학', 4, 2, 'AI 검토: 정답 오류 확인', '원본 시험지 확인 후 DB 수정'],
    [34, 'ELEC-E-EM-411',  '2024년', '3회', 'EM-11', '전기자기학',         3, 1, 'AI 검토: 정답 오류 확인', '원본 시험지 확인 후 DB 수정'],
    [35, 'ELEC-E-EM-419',  '2024년', '3회', 'EM-19', '전기자기학',         2, 3, 'AI 검토: 정답 오류 확인', '원본 시험지 확인 후 DB 수정'],
    [36, 'ELEC-E-EM-420',  '2024년', '3회', 'EM-20', '전기자기학',         2, 3, 'AI 검토: 정답 오류 확인', '원본 시험지 확인 후 DB 수정'],
    [37, 'ELEC-E-MA-453',  '2024년', '3회', 'MA-13', '전기기기',           2, 3, 'AI 검토: 정답 오류 확인', '원본 시험지 확인 후 DB 수정'],
    [38, 'ELEC-E-PC-433',  '2024년', '3회', 'PC-13', '전력공학',           1, 4, 'AI 검토: 정답 오류 확인', '원본 시험지 확인 후 DB 수정'],
    [39, 'ELEC-E-PC-437',  '2024년', '3회', 'PC-17', '전력공학',           1, 2, 'AI 검토: 정답 오류 확인', '원본 시험지 확인 후 DB 수정'],
    [40, 'ELEC-E-PC-439',  '2024년', '3회', 'PC-19', '전력공학',           3, 1, 'AI 검토: 정답 오류 확인', '원본 시험지 확인 후 DB 수정'],
    # --- 2025년 ---
    [41, 'ELEC-E-CC-297',  '2025년', '3회', 77,     '회로이론 및 제어공학', 3, 1, '라플라스 역변환 결과 e^(-t)·sin(2t)이 정답. 정답 ①. DB에 ③으로 저장.', 'DB 정답 3→1 수정 필요'],
]


# ============================================================
# 2. 규정 변경 (2건)
# ============================================================
# [No, 코드, 년도, 회차, 문제번호, 과목, DB정답, 현행정답, 오류설명, 조치사항]
reg_changes = [
    [1, 'ELEC-E-EL-036', '2018년', '2회', 96,     '전기설비기술기준', 1, 2, '누전차단기 감도전류가 KEC 2021 개정으로 15mA→30mA. 구기준 정답 ①(15mA), 현행법 정답 ②(30mA).', 'DB 정답 1→2 수정하거나, 문제 비활성화(is_active=false) 처리'],
    [2, 'ELEC-E-EL-408', '2023년', '2회', 'EL-08', '전기설비기술기준', 2, '없음', 'KEC 2021 이후 누전차단기 감도전류 15mA→30mA. 보기 전부 구기준(15mA)이라 현행법 기준 정답 없음.', '문제 비활성화(is_active=false) 처리 권장'],
]


# ============================================================
# 3. OCR/텍스트 오류 (15건 = 기존 10 + 신규 5)
# ============================================================
# [No, 코드, 년도, 회차, 문제번호, 과목, 오류위치, 오류설명, 조치사항]
ocr_errors = [
    # --- 2018년 ---
    [1,  'ELEC-E-EL-016',  '2018년', '1회', 96,     '전기설비기술기준', '문제', '문제 텍스트에 특수문자 깨짐', '원본 대조 후 특수문자/기호 수정 필요'],
    [2,  'ELEC-E-XX-032',  '2018년', '1회', 52,     '전기기기',       '보기', '보기 기호 매칭 오류. 보기 1~4의 텍스트가 잘못된 칸에 배치', '원본 대조 후 보기 텍스트 재배치'],
    [3,  'ELEC-E-XX-057',  '2018년', '2회', 37,     '전력공학',       '보기', '보기에 kW가 KW로 표기. SI 단위 오류', 'DB에서 KW→kW 텍스트 수정'],
    # --- 2019년 ---
    [4,  'ELEC-E-EL-082',  '2019년', '2회', 82,     '전기설비기술기준', '문제', '보기 일부 누락. 첨자/분수 표현 불완전', '원본 대조 후 첨자/분수 표현 보완'],
    [5,  'ELEC-E-XX-143',  '2019년', '1회', 43,     '전기기기',       '보기', '수식 누락. 보기의 분수/첨자 텍스트가 완전하지 않음', '원본 대조 후 KaTeX 수식 재입력'],
    [6,  'ELEC-E-XX-196',  '2019년', '2회', 56,     '전기기기',       '문제', '문제 텍스트 일부가 중복 입력됨', 'DB에서 중복 텍스트 제거'],
    # --- 2020년 ---
    [7,  'ELEC-E-XX-244',  '2020년', '1회', 24,     '전력공학',       '보기', '표/그림의 숫자가 정확히 표현 안됨', '원본 대조 후 숫자 수정'],
    # --- 2021년 ---
    [8,  'ELEC-E-XX-450',  '2021년', '3회', 30,     '전력공학',       '문제', '문제 텍스트에 불필요한 줄바꿈이 포함되어 가독성 저해', 'DB에서 불필요한 줄바꿈 제거'],
    # --- 2023년 ---
    [9,  'ELEC-E-CC-381',  '2023년', '2회', 'CC-01', '회로이론 및 제어공학', '보기', '보기4 "(s+3)" → "(s+3)(s+5)" 누락. 원본 확인 완료', '수정완료 (DB 반영됨)'],
    [10, 'ELEC-E-CC-394',  '2023년', '2회', 'CC-14', '회로이론 및 제어공학', '문제', '전달함수 분자 "s" → "10"(상수). 원본 확인 완료', '수정완료 (DB 반영됨)'],
    [11, 'ELEC-E-CC-399',  '2023년', '2회', 'CC-19', '회로이론 및 제어공학', '보기', '보기4 s^(n+1) → s^n. 원본 확인 완료', '수정완료 (DB 반영됨)'],
    [12, 'ELEC-E-PC-375',  '2023년', '3회', 'PC-15', '전력공학',           '보기', '보기1 "지상전류": 원본에도 "지상전류"로 인쇄. OCR 정확하나 출제측 오류 가능', '조치불필요 (원본과 일치)'],
    # --- 2024년 ---
    [13, 'ELEC-E-CC-466',  '2024년', '3회', 'CC-06', '회로이론 및 제어공학', '문제', '숫자값 원본과 일치하나 어떤 보기도 계산결과와 불일치', '조치불필요 (원본과 일치, 출제 오류 추정)'],
    # --- 2025년 ---
    [14, 'ELEC-E-MA-351',  '2025년', '2회', 51,     '전기기기',       '보기', '보기 3, 4번의 텍스트가 잘려 있음', '원본 대조 후 보기 텍스트 보완'],
    [15, 'ELEC-E-MA-359',  '2025년', '2회', 59,     '전기기기',       '보기', '수식 레이아웃 깨짐. KaTeX 수식이 올바르게 렌더링 안 됨', '원본 대조 후 수식 재입력'],
]


# ============================================================
# 시트 1: 정답 오류
# ============================================================
ws_err = wb.active
ws_err.title = '정답 오류 (41건)'
setup_sheet(ws_err,
    ['No', '문제코드', '년도', '회차', '문제번호', '과목', 'DB정답', '올바른 정답', '오류 설명', '조치사항'],
    answer_errors,
    [5, 20, 8, 6, 12, 22, 8, 10, 55, 30],
    header_fill_blue,
    action_col=10
)

# ============================================================
# 시트 2: 규정 변경
# ============================================================
ws_reg = wb.create_sheet('규정 변경 (2건)')
setup_sheet(ws_reg,
    ['No', '문제코드', '년도', '회차', '문제번호', '과목', 'DB정답', '현행 정답', '오류 설명', '조치사항'],
    reg_changes,
    [5, 20, 8, 6, 12, 22, 8, 10, 55, 35],
    header_fill_orange,
    action_col=10
)

# ============================================================
# 시트 3: OCR/텍스트 오류
# ============================================================
ws_ocr = wb.create_sheet('OCR·텍스트 오류 (15건)')
setup_sheet(ws_ocr,
    ['No', '문제코드', '년도', '회차', '문제번호', '과목', '오류 위치', '오류 설명', '조치사항'],
    ocr_errors,
    [5, 20, 8, 6, 12, 22, 10, 55, 30],
    header_fill_green,
    action_col=9
)

# ============================================================
# 시트 0: 요약
# ============================================================
ws_sum = wb.create_sheet('요약', 0)
summary = [
    ['전기기사 필기 문제 검토 오류 목록'],
    [],
    ['검토 범위', '전기기사 필기 2018~2025년 전체 (21개 시험, 2,100문제)'],
    ['검토 일자', '2026-04-06'],
    [],
    ['구분', '건수', '상태'],
    ['정답 오류', '41건', 'DB정답 수정 필요 32건 / 원본 대조 필요 9건'],
    ['규정 변경', '2건', '문제 비활성화 또는 정답 변경 필요'],
    ['OCR·텍스트 오류', '15건', 'DB수정 완료 3건 / 원본 대조 필요 7건 / 조치불필요 2건 / 텍스트 수정 필요 3건'],
    ['합계', '58건', ''],
    [],
    ['조치사항 범례'],
    ['배경색: 초록', '수정 완료 또는 조치 불필요'],
    ['배경색: 노랑', 'DB 수정 필요 (정답 변경 등)'],
    ['배경색: 빨강', '원본 시험지 대조 확인 후 DB 수정'],
    [],
    ['참고사항'],
    ['- 정답 오류 32~40번(2024년 3회, 9건)은 AI 검토 결과이며, 원본 시험지와 대조 확인 후 DB 수정 필요'],
    ['- 규정 변경 문제는 KEC 2021 개정 관련. 현행 기준에 맞지 않으므로 비활성화 권장'],
    ['- OCR 오류 중 "원본 대조" 항목은 원본 시험지 PDF/이미지를 직접 확인해야 함'],
]

for i, row in enumerate(summary, 1):
    for j, val in enumerate(row, 1):
        cell = ws_sum.cell(row=i, column=j, value=val)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        if i == 1:
            cell.font = Font(bold=True, size=14)
        elif i == 6:
            cell.font = header_font
            cell.fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif val and any(k in str(val) for k in ['검토 범위', '합계', '조치사항 범례', '참고사항']):
            cell.font = Font(bold=True, size=11)
        elif '배경색: 초록' in str(val) if val else False:
            cell.fill = done_fill
        elif '배경색: 노랑' in str(val) if val else False:
            cell.fill = todo_fill
        elif '배경색: 빨강' in str(val) if val else False:
            cell.fill = verify_fill

ws_sum.column_dimensions['A'].width = 25
ws_sum.column_dimensions['B'].width = 15
ws_sum.column_dimensions['C'].width = 55

try:
    wb.save('data/전기기사_문제검토_오류목록.xlsx')
except PermissionError:
    wb.save('data/전기기사_문제검토_오류목록_updated.xlsx')
    print('  (원본 파일이 열려 있어 _updated 파일로 저장)')

err_todo = sum(1 for r in answer_errors if 'DB 정답' in r[9])
err_verify = sum(1 for r in answer_errors if '원본 시험지' in r[9])
ocr_done = sum(1 for r in ocr_errors if '수정완료' in r[8])
ocr_noaction = sum(1 for r in ocr_errors if '조치불필요' in r[8])
ocr_todo = len(ocr_errors) - ocr_done - ocr_noaction

print('Excel 저장 완료')
print(f'  정답 오류: {len(answer_errors)}건 (DB수정 필요 {err_todo}건, 원본대조 {err_verify}건)')
print(f'  규정 변경: {len(reg_changes)}건')
print(f'  OCR 오류: {len(ocr_errors)}건 (완료 {ocr_done}건, 조치필요 {ocr_todo}건, 불필요 {ocr_noaction}건)')
print(f'  총합계: {len(answer_errors)+len(reg_changes)+len(ocr_errors)}건')
